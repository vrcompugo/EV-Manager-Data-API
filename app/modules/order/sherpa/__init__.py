import os
import tempfile
from openpyxl import load_workbook, Workbook

from app.models import Order


def generate_sherpa_file(deal, order: Order):
    from app.modules.order.sherpa.excel.part02_rechnungsanschrift import part02_rechnungsanschrift
    from app.modules.order.sherpa.excel.part03_bankdaten import part03_bankdaten
    from app.modules.order.sherpa.excel.part03_bankdaten_inhaber import part03_bankdaten_inhaber
    from app.modules.order.sherpa.excel.part05_lieferstelle import part05_lieferstelle
    from app.modules.order.sherpa.excel.part06_zahlerdaten import part06_zahlerdaten
    from app.modules.order.sherpa.excel.part07_marktpartner import part07_marktpartner
    from app.modules.order.sherpa.excel.part08_weitere_daten import part08_weitere_daten
    from app.modules.order.sherpa.excel.part11_sonstige_daten import part11_sonstige_daten
    from app.modules.order.sherpa.excel.part12_stammdaten_tarif import part12_stammdaten_tarif

    book = Workbook()
    excel_layout = "cloud"
    if order.contract_number is not None:
        if order.contract_number[0:1] == "G":
            excel_layout = "gas"
        if order.contract_number[0:1] == "E":
            excel_layout = "power"
    wb = load_workbook(os.path.join(os.path.dirname(__file__), f'excel/basefiles/{excel_layout}.xlsx'))
    wb.current_row = 4
    part02_rechnungsanschrift(wb, excel_layout, order)
    part03_bankdaten(wb, excel_layout, order)
    part03_bankdaten_inhaber(wb, excel_layout, order)
    part05_lieferstelle(wb, excel_layout, order)
    part06_zahlerdaten(wb, excel_layout, order)
    part07_marktpartner(wb, excel_layout, order)
    part08_weitere_daten(wb, excel_layout, order)
    part11_sonstige_daten(wb, excel_layout, order)
    part12_stammdaten_tarif(wb, excel_layout, order)
    with tempfile.TemporaryDirectory() as tmpdirname:
        wb.save(filename=f"{tmpdirname}/workbook.xlsx")
        with open(f"{tmpdirname}/workbook.xlsx", "rb") as f:
            filecontent = f.read()

    return filecontent


def generate_sherpa2_file(deal, contact):
    from app.modules.order.sherpa.excel2.part02_rechnungsanschrift import part02_rechnungsanschrift
    from app.modules.order.sherpa.excel2.part03_bankdaten import part03_bankdaten
    from app.modules.order.sherpa.excel2.part03_bankdaten_inhaber import part03_bankdaten_inhaber
    from app.modules.order.sherpa.excel2.part05_lieferstelle import part05_lieferstelle
    from app.modules.order.sherpa.excel2.part06_zahlerdaten import part06_zahlerdaten
    from app.modules.order.sherpa.excel2.part07_marktpartner import part07_marktpartner
    from app.modules.order.sherpa.excel2.part08_weitere_daten import part08_weitere_daten
    from app.modules.order.sherpa.excel2.part11_sonstige_daten import part11_sonstige_daten
    from app.modules.order.sherpa.excel2.part12_stammdaten_tarif import part12_stammdaten_tarif

    book = Workbook()
    excel_layout = "cloud"
    if deal.get("contract_number") is not None:
        if deal.get("contract_number")[0:1] == "G":
            excel_layout = "gas"
        if deal.get("contract_number")[0:1] == "E":
            excel_layout = "power"
    wb = load_workbook(os.path.join(os.path.dirname(__file__), f'excel/basefiles/{excel_layout}.xlsx'))
    wb.current_row = 4
    part02_rechnungsanschrift(wb, excel_layout, deal, contact)
    part03_bankdaten(wb, excel_layout, deal, contact)
    part03_bankdaten_inhaber(wb, excel_layout, deal, contact)
    part05_lieferstelle(wb, excel_layout, deal, contact)
    part06_zahlerdaten(wb, excel_layout, deal, contact)
    part07_marktpartner(wb, excel_layout, deal, contact)
    part08_weitere_daten(wb, excel_layout, deal, contact)
    part11_sonstige_daten(wb, excel_layout, deal, contact)
    part12_stammdaten_tarif(wb, excel_layout, deal, contact)
    with tempfile.TemporaryDirectory() as tmpdirname:
        wb.save(filename=f"{tmpdirname}/workbook.xlsx")
        with open(f"{tmpdirname}/workbook.xlsx", "rb") as f:
            filecontent = f.read()

    return filecontent
